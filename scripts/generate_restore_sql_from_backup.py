import argparse
import json
from pathlib import Path

from openpyxl import load_workbook


RESTORE_ORDER = [
    "profiles",
    "customers",
    "suppliers",
    "inv_categories",
    "inv_products",
    "inv_variants",
    "inv_rolls",
    "inv_movements",
    "fg_stock",
    "product_codes",
    "product_recipes",
    "recipe_items",
    "rrp_entries",
    "orders",
    "order_items",
    "order_components",
    "wastage_logs",
    "execution_logs",
    "activity_logs",
    "order_tickets",
    "order_ticket_followups",
    "order_quote_forms",
    "order_quote_downloads",
    "stock_orders",
    "stock_order_items",
    "stock_order_downloads",
]

JSON_COLUMNS = {
    ("activity_logs", "changes"),
    ("rrp_entries", "price_map"),
    ("order_quote_forms", "form_data"),
    ("order_quote_downloads", "form_data"),
    ("stock_orders", "order_form_data"),
    ("stock_order_downloads", "form_data"),
}

GENERATED_COLUMNS = {
    ("wastage_logs", "waste_length_m"),
    ("wastage_logs", "waste_width_m"),
    ("wastage_logs", "waste_area_sqm"),
}

PRIMARY_KEYS = {
    "order_quote_forms": "order_id",
}


def quote_ident(name):
    return '"' + name.replace('"', '""') + '"'


def dollar_json(value):
    text = json.dumps(value, ensure_ascii=False, separators=(",", ":"))
    return f"$restore_json${text}$restore_json$::jsonb"


def normalize_value(table, column, value):
    if value is None or value == "":
        return None
    if (table, column) in JSON_COLUMNS:
        if isinstance(value, str):
            try:
                return json.loads(value)
            except json.JSONDecodeError:
                return value
        return value
    return value


def read_sheet_rows(workbook, table):
    ws = workbook[table]
    rows_iter = ws.iter_rows(values_only=True)
    header = next(rows_iter, None)
    if not header or header[0] == "No exported rows":
        return [], []

    source_columns = [str(col) for col in header if col is not None]
    columns = [col for col in source_columns if (table, col) not in GENERATED_COLUMNS]
    rows = []

    for raw in rows_iter:
        row = {}
        has_value = False
        for index, source_column in enumerate(source_columns):
            if (table, source_column) in GENERATED_COLUMNS:
                continue
            value = raw[index] if index < len(raw) else None
            normalized = normalize_value(table, source_column, value)
            if normalized is not None:
                has_value = True
            row[source_column] = normalized
        if has_value:
            rows.append(row)

    return columns, rows


def max_ticket_number(workbook):
    if "order_tickets" not in workbook.sheetnames:
        return 0

    rows = list(workbook["order_tickets"].iter_rows(values_only=True))
    if not rows or not rows[0] or rows[0][0] == "No exported rows":
        return 0

    headers = [str(col) for col in rows[0]]
    if "ticket_uid" not in headers:
        return 0

    index = headers.index("ticket_uid")
    values = []
    for row in rows[1:]:
        value = row[index]
        if value is not None and str(value).isdigit():
            values.append(int(str(value)))
    return max(values) if values else 0


def build_restore_sql(backup_path):
    workbook = load_workbook(backup_path, read_only=True, data_only=True)
    missing = [table for table in RESTORE_ORDER if table not in workbook.sheetnames]
    if missing:
        raise RuntimeError(f"Backup is missing expected sheet(s): {', '.join(missing)}")

    truncate_tables = [table for table in RESTORE_ORDER if table != "profiles"]
    counts = {}
    parts = [
        f"-- Generated restore SQL from {backup_path.as_posix()}",
        "-- Restores public app tables. Auth users are not modified.",
        "-- Generated columns from backup are skipped so Postgres recomputes them.",
        "BEGIN;",
        "",
        "TRUNCATE TABLE",
        ",\n".join(f"  public.{quote_ident(table)}" for table in truncate_tables),
        "RESTART IDENTITY CASCADE;",
        "",
    ]

    for table in RESTORE_ORDER:
        columns, rows = read_sheet_rows(workbook, table)
        counts[table] = len(rows)
        if not rows:
            continue

        pk = PRIMARY_KEYS.get(table, "id")
        column_list = ", ".join(quote_ident(column) for column in columns)
        select_list = ", ".join(quote_ident(column) for column in columns)
        updates = ", ".join(
            f"{quote_ident(column)} = EXCLUDED.{quote_ident(column)}"
            for column in columns
            if column != pk
        )
        conflict_sql = (
            f"ON CONFLICT ({quote_ident(pk)}) DO UPDATE SET {updates}"
            if updates
            else f"ON CONFLICT ({quote_ident(pk)}) DO NOTHING"
        )

        for start in range(0, len(rows), 150):
            chunk = rows[start:start + 150]
            parts.extend([
                f"-- {table}: rows {start + 1}-{start + len(chunk)} of {len(rows)}",
                f"INSERT INTO public.{quote_ident(table)} ({column_list})",
                f"SELECT {select_list}",
                f"FROM jsonb_populate_recordset(NULL::public.{quote_ident(table)}, {dollar_json(chunk)}) AS r",
                f"{conflict_sql};",
                "",
            ])

    max_ticket = max_ticket_number(workbook)
    parts.extend([
        "DO $$",
        "BEGIN",
        "  IF to_regclass('public.order_ticket_number_seq') IS NOT NULL THEN",
        (
            f"    PERFORM setval('public.order_ticket_number_seq', {max_ticket}, true);"
            if max_ticket
            else "    PERFORM setval('public.order_ticket_number_seq', 1, false);"
        ),
        "  END IF;",
        "END $$;",
        "",
        "COMMIT;",
        "",
        "NOTIFY pgrst, 'reload schema';",
        "",
        "-- Expected restored row counts:",
    ])

    for table in RESTORE_ORDER:
        parts.append(f"-- {table}: {counts.get(table, 0)}")
    parts.append(f"-- ticket_sequence_max: {max_ticket}")

    return "\n".join(parts), counts, max_ticket


def main():
    parser = argparse.ArgumentParser(description="Generate Supabase restore SQL from a Vista backup workbook.")
    parser.add_argument("--backup", required=True, help="Path to vista_supabase_backup_*.xlsx")
    parser.add_argument("--output", required=True, help="Path for generated restore SQL")
    args = parser.parse_args()

    backup_path = Path(args.backup)
    output_path = Path(args.output)
    sql, counts, max_ticket = build_restore_sql(backup_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(sql, encoding="utf-8")

    print(output_path)
    for table in RESTORE_ORDER:
        print(f"{table}: {counts.get(table, 0)}")
    print(f"ticket_sequence_max: {max_ticket}")


if __name__ == "__main__":
    main()
